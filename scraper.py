from __future__ import annotations

import csv
import html
import json
import logging
import os
import re
import shutil
import sys
import time
from dataclasses import asdict, dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable
from urllib.parse import urljoin, urlparse, urlunparse

import requests
from bs4 import BeautifulSoup, Tag
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

SCRAPER_VERSION = "2026-07-17-v7-stock-filter"

BASE_URL = "https://akvasport.com/"
CATEGORY_URL = os.getenv(
    "CATEGORY_URL", "https://akvasport.com/category/291/primamki.html"
).strip()
PRODUCT_LIMIT = int(os.getenv("PRODUCT_LIMIT", "10"))
REQUEST_DELAY = float(os.getenv("REQUEST_DELAY", "1.2"))
DEFAULT_IN_STOCK_QTY = int(os.getenv("DEFAULT_IN_STOCK_QTY", "1"))
TEMPLATE_PATH = Path(os.getenv("TEMPLATE_PATH", "input/TEMU_TEMPLATE.xlsx"))
OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", "output"))
OUTPUT_XLSX = OUTPUT_DIR / "TEMU_AKVASPORT_UPLOAD.xlsx"
OUTPUT_CSV = OUTPUT_DIR / "akvasport_raw_export.csv"
OUT_OF_STOCK_CSV = OUTPUT_DIR / "akvasport_out_of_stock.csv"
REVIEW_CSV = OUTPUT_DIR / "akvasport_needs_review.csv"
LOG_PATH = OUTPUT_DIR / "scraper.log"
INCLUDE_OUT_OF_STOCK = os.getenv("INCLUDE_OUT_OF_STOCK", "0").strip().lower() in {"1", "true", "yes", "y"}

MANUFACTURER = "Tianyun Fishing Tackle Co.,Ltd"
EU_RESPONSIBLE_PERSON = "AKVASPORT EOOD"
SHIPPING_TEMPLATE = "Магазин"
HANDLING_TIME = "1 Day"
FULFILLMENT_CHANNEL = "I will ship this item myself"
DEFAULT_COUNTRY_OF_ORIGIN = os.getenv("DEFAULT_COUNTRY_OF_ORIGIN", "Mainland China").strip()

CATEGORY_NAMES = {
    "32474": "Sports & Outdoors / Hunting & Fishing / Fishing / Baits & Accessories / Bait Traps",
    "32476": "Sports & Outdoors / Hunting & Fishing / Fishing / Baits & Accessories / Baits & Attractants / Artificial Bait",
    "32477": "Sports & Outdoors / Hunting & Fishing / Fishing / Baits & Accessories / Baits & Attractants / Attractants",
    "32478": "Sports & Outdoors / Hunting & Fishing / Fishing / Baits & Accessories / Baits & Attractants / Eggs",
    "32479": "Sports & Outdoors / Hunting & Fishing / Fishing / Baits & Accessories / Baits & Attractants / Light Attractants",
}

# Template columns. Row 2 contains the user-facing labels.
COL = {
    "category": "E",
    "category_name": "F",
    "product_type": "G",
    "product_name": "L",
    "contribution_goods": "M",
    "contribution_sku": "N",
    "update_or_add": "O",
    "brand": "R",
    "description": "T",
    "bullet_1": "U",
    "bullet_2": "V",
    "bullet_3": "W",
    "bullet_4": "X",
    "bullet_5": "Y",
    "detail_img_start": "AA",
    "major_material_1": "DT",
    "major_material_2": "DU",
    "major_material_3": "DV",
    "power_mode": "EL",
    "variation_theme": "ET",
    "color": "EU",
    "size": "EV",
    "style": "EW",
    "material": "EX",
    "flavor": "EY",
    "capacity": "FA",
    "weight_variant": "FC",
    "items": "FD",
    "quantity_variant": "FE",
    "model": "FF",
    "sku_img_start": "FH",
    "quantity": "FR",
    "base_price": "FS",
    "reference_link": "FT",
    "list_price": "FU",
    "no_list_price": "FV",
    "weight_g": "FW",
    "length_cm": "FX",
    "width_cm": "FY",
    "height_cm": "FZ",
    "sku_type": "GA",
    "individually_packed": "GB",
    "total_packaging_qty": "GC",
    "packaging_unit": "GD",
    "shipping_template": "GL",
    "handling_time": "GM",
    "fulfillment": "GN",
    "country_origin": "GP",
    "product_identification": "IQ",
    "manufacturer": "IR",
    "eu_responsible": "IS",
}

REQUIRED_BY_CATEGORY = {
    cat: {
        "E", "L", "M", "N", "DT", "DU", "DV", "ET", "FH", "FR", "FS",
        "FU", "FW", "FX", "FY", "FZ", "GB", "GC", "GD", "GL", "GP",
        "IQ", "IR", "IS",
    }
    for cat in CATEGORY_NAMES
}
REQUIRED_BY_CATEGORY["32479"] = (
    REQUIRED_BY_CATEGORY["32479"] - {"DT", "DU", "DV"}
) | {"EL"}

logger = logging.getLogger("akvasport")


@dataclass
class Variant:
    sku: str
    option_values: dict[str, str] = field(default_factory=dict)
    price_eur: Decimal = Decimal("0")
    quantity: int = 0
    weight_g: float | None = None
    image_urls: list[str] = field(default_factory=list)


@dataclass
class Product:
    url: str
    title: str
    parent_sku: str
    brand: str
    description: str
    bullets: list[str]
    category_id: str
    detail_images: list[str]
    attributes: dict[str, str]
    list_price_eur: Decimal | None
    country_origin: str
    country_origin_source: str
    variants: list[Variant]


def setup_logging() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    logger.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
    stream = logging.StreamHandler(sys.stdout)
    stream.setFormatter(formatter)
    file_handler = logging.FileHandler(LOG_PATH, encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.handlers[:] = [stream, file_handler]


def build_session() -> requests.Session:
    retry = Retry(
        total=5,
        connect=5,
        read=5,
        backoff_factor=1.0,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset(("GET", "POST")),
        raise_on_status=False,
    )
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/136.0 Safari/537.36"
            ),
            "Accept-Language": "bg-BG,bg;q=0.9,en;q=0.7",
        }
    )
    session.mount("https://", HTTPAdapter(max_retries=retry))
    session.mount("http://", HTTPAdapter(max_retries=retry))
    return session


def fetch(session: requests.Session, url: str, *, method: str = "GET", **kwargs) -> requests.Response:
    time.sleep(max(0.0, REQUEST_DELAY))
    response = session.request(method, url, timeout=45, **kwargs)
    response.raise_for_status()
    response.encoding = response.apparent_encoding or response.encoding or "utf-8"
    return response


def clean_text(value: str | None) -> str:
    if not value:
        return ""
    value = html.unescape(value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def normalized_url(url: str) -> str:
    absolute = urljoin(BASE_URL, url)
    parsed = urlparse(absolute)
    return urlunparse(("https", parsed.netloc.lower(), parsed.path, "", "", ""))


def normalized_page_url(url: str) -> str:
    """Normalize category/pagination URLs while preserving ?page=N."""
    absolute = urljoin(BASE_URL, url)
    parsed = urlparse(absolute)
    return urlunparse(
        ("https", parsed.netloc.lower(), parsed.path, "", parsed.query, "")
    )


def original_image_url(url: str) -> str:
    url = normalized_url(url)
    replacements = (
        (".thumb.webp", ".webp"),
        (".box.webp", ".webp"),
        (".large.webp", ".webp"),
        (".thumb.jpg", ".jpg"),
        (".box.jpg", ".jpg"),
        (".large.jpg", ".jpg"),
        (".thumb.jpeg", ".jpeg"),
        (".box.jpeg", ".jpeg"),
        (".large.jpeg", ".jpeg"),
        (".thumb.png", ".png"),
        (".box.png", ".png"),
        (".large.png", ".png"),
    )
    for old, new in replacements:
        url = url.replace(old, new)
    return url


def unique(items: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        if not item or item in seen:
            continue
        seen.add(item)
        result.append(item)
    return result


def collect_product_urls(
    session: requests.Session, category_url: str, limit: int
) -> list[str]:
    """Collect product URLs from every category page.

    Product URLs are normalized without query parameters, while category page
    URLs preserve ?page=N. This distinction is essential for pagination.
    """
    urls: list[str] = []
    seen_pages: set[str] = set()
    page_url = normalized_page_url(category_url)

    while page_url and page_url not in seen_pages:
        seen_pages.add(page_url)
        logger.info("Category page: %s", page_url)
        soup = BeautifulSoup(fetch(session, page_url).text, "lxml")

        page_products: list[str] = []
        for anchor in soup.select('a[href*="/product/"]'):
            href = anchor.get("href")
            if not href:
                continue
            product_url = normalized_url(href)
            if re.search(r"/product/\d+/", product_url):
                page_products.append(product_url)

        for product_url in unique(page_products):
            if product_url not in urls:
                urls.append(product_url)
                if limit > 0 and len(urls) >= limit:
                    return urls

        next_link = (
            soup.select_one('link[rel~="next"][href]')
            or soup.select_one('a[rel~="next"][href]')
        )

        if next_link:
            next_url = normalized_page_url(next_link.get("href", ""))
            page_url = next_url if next_url not in seen_pages else ""
            continue

        current_match = re.search(r"(?:\?|&)page=(\d+)", page_url)
        current_page = int(current_match.group(1)) if current_match else 1
        candidates: list[tuple[int, str]] = []

        for anchor in soup.select('a[href*="page="]'):
            href = anchor.get("href")
            if not href:
                continue
            candidate_url = normalized_page_url(href)
            page_match = re.search(r"(?:\?|&)page=(\d+)", candidate_url)
            if not page_match:
                continue
            page_number = int(page_match.group(1))
            if page_number > current_page and candidate_url not in seen_pages:
                candidates.append((page_number, candidate_url))

        page_url = min(candidates, default=(0, ""))[1]

    return urls

def parse_decimal(value: str | float | int | None) -> Decimal | None:
    if value is None:
        return None
    cleaned = re.sub(r"[^0-9,.-]", "", str(value)).replace(",", ".")
    try:
        return Decimal(cleaned).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def extract_json_assignment(page: str, variable: str) -> dict:
    pattern = rf"{re.escape(variable)}\s*=\s*(\{{.*?\}})\s*;"
    match = re.search(pattern, page, flags=re.DOTALL)
    if not match:
        return {}
    try:
        return json.loads(match.group(1))
    except json.JSONDecodeError:
        logger.warning("Could not decode %s", variable)
        return {}


def parse_attributes(soup: BeautifulSoup) -> dict[str, str]:
    result: dict[str, str] = {}
    for item in soup.select(".c-product-page__attribute-item"):
        name = clean_text(
            item.select_one(".c-product-page__attribute-name").get_text(" ", strip=True)
            if item.select_one(".c-product-page__attribute-name")
            else ""
        ).rstrip(":")
        value_node = item.select_one(".c-product-page__attribute-value")
        value = clean_text(value_node.get_text(" ", strip=True) if value_node else "")
        if name and value:
            result[name] = value
    return result


def extract_description(soup: BeautifulSoup) -> str:
    node = soup.select_one("#product-detailed-description")
    if not node:
        meta = soup.select_one('meta[property="og:description"]')
        return clean_text(meta.get("content") if meta else "")
    clone = BeautifulSoup(str(node), "lxml")
    for bad in clone.select("img, script, style, iframe, video, source, noscript"):
        bad.decompose()
    text = clone.get_text("\n", strip=True)
    lines = unique(clean_text(line) for line in text.splitlines() if clean_text(line))
    return "\n".join(lines)[:5000]


def extract_images(soup: BeautifulSoup) -> tuple[list[str], list[tuple[str, str]]]:
    gallery = soup.select_one("#product-images")
    if not gallery:
        gallery = soup
    images: list[tuple[str, str]] = []
    for anchor in gallery.select("a[href], a[ref]"):
        raw = anchor.get("ref") or anchor.get("href") or ""
        if "/userfiles/productimages/" not in raw:
            continue
        label = clean_text(anchor.get("title"))
        image = anchor.select_one("img")
        if image:
            label = clean_text(f"{label} {image.get('alt', '')}")
        images.append((original_image_url(raw), label))
    for image in gallery.select("img[src]"):
        raw = image.get("data-pinch-zoom-src") or image.get("src") or ""
        if "/userfiles/productimages/" not in raw:
            continue
        images.append((original_image_url(raw), clean_text(image.get("alt"))))
    dedup: list[tuple[str, str]] = []
    seen: set[str] = set()
    for url, label in images:
        if url not in seen:
            seen.add(url)
            dedup.append((url, label))
    return [u for u, _ in dedup], dedup


def option_groups(soup: BeautifulSoup) -> tuple[dict[str, dict[str, str]], list[tuple[str, str]]]:
    groups: dict[str, dict[str, str]] = {}
    ordered_groups: list[tuple[str, str]] = []
    for select in soup.select('select[name^="Options"], select.optionFilter'):
        name = select.get("name", "")
        data_type = select.get("data-type", "")
        group_id = select.get("data-id", "") or select.get("id", "").replace("optionGroup_", "")
        match = re.search(r"Options\]\[([^]]+)\]\[(\d+)\]", name)
        if match:
            data_type, group_id = match.group(1), match.group(2)
        if not group_id:
            continue
        label_node = select.find_previous(class_=re.compile("option-name|filter-title"))
        label = clean_text(label_node.get_text(" ", strip=True) if label_node else "Option").rstrip(":")
        values = {
            str(opt.get("value")): clean_text(opt.get_text(" ", strip=True))
            for opt in select.select("option[value]")
            if str(opt.get("value")) not in {"", "0"}
        }
        if values:
            groups[group_id] = values
            ordered_groups.append((group_id, label))
    return groups, unique_pairs(ordered_groups)


def unique_pairs(items: Iterable[tuple[str, str]]) -> list[tuple[str, str]]:
    seen: set[str] = set()
    out: list[tuple[str, str]] = []
    for key, value in items:
        if key in seen:
            continue
        seen.add(key)
        out.append((key, value))
    return out


def variant_option_values(
    variant_key: str,
    groups: dict[str, dict[str, str]],
    ordered_groups: list[tuple[str, str]],
) -> dict[str, str]:
    result: dict[str, str] = {}
    ids = re.findall(r"\d+", variant_key)
    used: set[str] = set()
    for option_id in ids:
        for group_id, label in ordered_groups:
            if option_id in groups.get(group_id, {}) and option_id not in used:
                result[label] = groups[group_id][option_id]
                used.add(option_id)
                break
    return result


def match_variant_images(
    option_values: dict[str, str],
    image_pairs: list[tuple[str, str]],
    all_images: list[str],
) -> list[str]:
    values = [v.lower() for v in option_values.values() if v]
    matched = [
        url
        for url, label in image_pairs
        if values and any(re.search(rf"(^|\W){re.escape(v)}($|\W)", label.lower()) for v in values)
    ]
    neutral = [
        url
        for url, label in image_pairs
        if not any(re.search(rf"(^|\W){re.escape(v)}($|\W)", label.lower()) for v in values)
    ]
    return unique(matched + neutral + all_images)[:10]


def parse_list_price(soup: BeautifulSoup, base_price: Decimal | None) -> Decimal | None:
    selectors = (
        ".old-price .price",
        ".list-price .price",
        ".u-price__old__value",
        "[data-list-price]",
    )
    for selector in selectors:
        for node in soup.select(selector):
            raw = node.get("data-list-price") or node.get_text(" ", strip=True)
            candidate = parse_decimal(raw)
            if candidate and base_price and candidate > base_price:
                return candidate
    return None


def infer_brand(title: str) -> str:
    known = [
        "Yo Zuri", "Yo-Zuri", "YO-ZURI", "Duel", "DUEL", "Owner", "Ragot",
        "Xesta", "SeaBuzz", "Filstar", "Shimano", "Daiwa", "Rapala", "Westin",
        "Mepps", "Blue Fox", "Storm", "Salmo", "DTD", "Shout", "Williamson",
        "Bait Breath", "Little Jack", "Seafloor Control", "ZetZ", "Mustad", "VMC",
        "Hayabusa", "Acme", "Goldy",
    ]
    lower = title.lower()
    for brand in known:
        if brand.lower() in lower:
            return "Yo-Zuri" if brand.lower() in {"yo zuri", "yo-zuri"} else brand
    first = re.split(r"\s+", title.strip())[0] if title.strip() else ""
    return first[:80]


def classify_category(title: str, description: str) -> str:
    # Classification is deliberately based mainly on the product title.
    # Words such as "light", "aroma" or "scent" inside a lure description must not
    # move a physical lure into Light Attractants or Attractants.
    title_text = clean_text(title).lower()

    if re.search(r"(?:bait\s*)?(?:trap|cage)|капан|капани|кош за стръв", title_text):
        return "32474"
    if re.search(r"\b(?:egg|eggs|roe)\b|яйц|хайвер|икра", title_text):
        return "32478"

    light_device = (
        re.search(r"(?:fish|fishing|bait|риболов|риба).{0,35}(?:led|lamp|light|лампа|светлин)", title_text)
        or re.search(r"(?:led|lamp|light|лампа|светлин).{0,35}(?:attract|привлич|риболов|риба)", title_text)
    )
    if light_device:
        return "32479"

    attractant_product = re.search(
        r"\b(?:attractant|dip|spray|booster|scent\s*(?:gel|spray|liquid)|flavou?r)\b"
        r"|атрактант|дип|спрей|бустер|ароматизатор",
        title_text,
    )
    if attractant_product:
        return "32477"

    return "32476"


COUNTRY_ALIASES = {
    "Japan": ("japan", "japanese", "япония", "японски", "японско"),
    "Mainland China": ("china", "made in china", "китай", "китайски"),
    "Bulgaria": ("bulgaria", "българия", "български"),
    "France": ("france", "франция", "френски"),
    "Italy": ("italy", "италия", "италиански"),
    "Germany": ("germany", "германия", "немски"),
    "Finland": ("finland", "финландия", "финландски"),
    "Sweden": ("sweden", "швеция", "шведски"),
    "Norway": ("norway", "норвегия", "норвежки"),
    "Denmark": ("denmark", "дания", "датски"),
    "Poland": ("poland", "полша", "полски"),
    "South Korea": ("south korea", "korea", "южна корея", "корея", "корейски"),
    "Taiwan": ("taiwan", "тайван", "тайвански"),
    "Vietnam": ("vietnam", "виетнам", "виетнамски"),
    "Thailand": ("thailand", "тайланд", "тайландски"),
    "Philippines": ("philippines", "филипини", "филипински"),
    "Malaysia": ("malaysia", "малайзия", "малайзийски"),
    "United States": ("united states", "usa", "u.s.a.", "сащ", "американски"),
}


def country_from_text(value: str, *, require_origin_context: bool = True) -> str | None:
    text = clean_text(value).lower()
    if not text:
        return None

    context_patterns = (
        r"made\s+in",
        r"manufactured\s+in",
        r"country\s+of\s+origin",
        r"произведен[а-я]*\s+в",
        r"производство",
        r"страна\s+на\s+произход",
    )
    windows: list[str] = []
    for pattern in context_patterns:
        for match in re.finditer(pattern, text):
            windows.append(text[match.start() : match.end() + 90])

    if require_origin_context and not windows:
        return None
    search_areas = windows or [text]

    for area in search_areas:
        for country, aliases in COUNTRY_ALIASES.items():
            if any(alias in area for alias in aliases):
                return country
    return None


def infer_country_origin(
    title: str, description: str, attributes: dict[str, str]
) -> tuple[str, str]:
    for name, value in attributes.items():
        if re.search(r"произход|country|origin|производство|manufactur", name, re.I):
            country = country_from_text(value, require_origin_context=False)
            if country:
                return country, f"attribute: {name}"

    country = country_from_text(description, require_origin_context=True)
    if country:
        return country, "explicit product description"

    # The field is mandatory in the Temu template. When AkvaSport does not publish
    # an explicit origin, use the configurable fallback and record this in the CSV/log.
    return DEFAULT_COUNTRY_OF_ORIGIN, "configured fallback"


MATERIAL_PATTERNS: list[tuple[str, str]] = [
    ("Tungsten Steel", r"\b(?:tungsten)\b|волфрам"),
    ("Lead", r"\blead\b|\bолово\b|\bоловна\s+глав|\bоловен\s+материал"),
    ("Stainless Steel", r"stainless\s+steel|неръждаема\s+стомана"),
    ("Carbon Steel", r"carbon\s+steel|въглеродна\s+стомана"),
    ("Silicone", r"\bsilicone\b|силикон"),
    ("Synthetic Rubber", r"synthetic\s+rubber|синтетичен\s+каучук"),
    ("Rubber", r"\brubber\b|\bкаучук\b"),
    ("PVC", r"\bpvc\b|поливинилхлорид"),
    ("ABS", r"\babs\b|abs\s+resin|abs\s+смол"),
    ("Resin", r"\bresin\b|\bсмола\b|смолен\s+материал"),
    ("Zinc Alloy", r"zinc\s+alloy|цинкова\s+сплав"),
    ("Copper Alloy", r"copper\s+alloy|медна\s+сплав"),
    ("Aluminum Alloy", r"alumin(?:um|ium)\s+alloy|алуминиева\s+сплав"),
    ("Iron", r"\biron\b|\bжелязо\b"),
    ("PC", r"polycarbonate|поликарбонат"),
    ("PE (polyethylene)", r"polyethylene|полиетилен"),
    ("PA (polyamide, Nylon)", r"polyamide|полиамид|\bnylon\b|\bнайлон\b"),
]


def material_values(
    title: str, description: str, attributes: dict[str, str], category_id: str
) -> tuple[str, str, str]:
    if category_id == "32479":
        return "", "", ""

    title_text = clean_text(title).lower()
    intro_text = clean_text(description)[:900].lower()
    attribute_text = " ".join(
        f"{name} {value}" for name, value in attributes.items()
    ).lower()
    evidence_text = f"{title_text} {intro_text} {attribute_text}"

    # First identify the actual product form. This prevents incidental words in the
    # description (for example, a tungsten sinker used with silicone lures) from
    # becoming the product's own material.
    if re.search(r"волфрам|\btungsten\b", title_text):
        return "Tungsten Steel", "Tungsten Steel", "Tungsten Steel"

    if re.search(r"джиг\s*глав|jig\s*head|micro\s*jig|глав[аи]\s+за\s+(?:туистер|силикон)", title_text):
        if re.search(r"волфрам|\btungsten\b", evidence_text):
            return "Tungsten Steel", "Carbon Steel", "Stainless Steel"
        return "Lead", "Carbon Steel", "Stainless Steel"

    if re.search(r"силикон|soft\s*bait|soft\s*lure|shad|туистер", title_text):
        if re.search(r"джиг\s*глав|jig\s*head", intro_text):
            return "Silicone", "Lead", "Carbon Steel"
        return "Silicone", "Silicone", "Silicone"

    # Hybrid lures may not say "silicone" in the title, but the first product
    # sentence clearly identifies both the soft body and integrated jig head.
    if re.search(r"силикон|soft\s*bait|soft\s*lure", intro_text) and re.search(
        r"джиг\s*глав|jig\s*head", intro_text
    ):
        return "Silicone", "Lead", "Carbon Steel"

    if re.search(r"воблер|wobbler|minnow|crankbait|jerkbait", title_text):
        return "ABS", "Resin", "Stainless Steel"

    if re.search(r"клатушка|блесна|пилкер|metal\s*jig|spinner|spoon", title_text):
        return "Stainless Steel", "Zinc Alloy", "Carbon Steel"

    materials: list[str] = []
    for material, pattern in MATERIAL_PATTERNS:
        if re.search(pattern, evidence_text, flags=re.I) and material not in materials:
            materials.append(material)

    if category_id == "32474" and not materials:
        materials = ["Stainless Steel", "PA (polyamide, Nylon)", "PE (polyethylene)"]
    elif category_id in {"32477", "32478"} and not materials:
        materials = ["Silicone", "PVC", "Rubber"]
    elif not materials:
        materials = ["ABS", "Stainless Steel", "Resin"]

    while len(materials) < 3:
        materials.append(materials[0])
    return tuple(materials[:3])


def parse_number(value: str) -> float | None:
    match = re.search(r"(\d+(?:[.,]\d+)?)", clean_text(value))
    return float(match.group(1).replace(",", ".")) if match else None


def parse_weight_g(value: str) -> float | None:
    text = clean_text(value).lower()
    number = parse_number(text)
    if number is None:
        return None
    if re.search(r"(?:kg)(?:\b|$)|килограм", text):
        return number * 1000
    if re.search(r"(?:mg)(?:\b|$)|милиграм", text):
        return number / 1000
    if re.search(r"(?:gr|g|гр)(?:\b|$)|грам", text):
        return number
    return None


def parse_length_cm(value: str) -> float | None:
    text = clean_text(value).lower()
    number = parse_number(text)
    if number is None:
        return None
    if re.search(r"(?:mm)(?:\b|$)|милимет", text):
        return number / 10
    if re.search(r"(?:cm|см)(?:\b|$)|сантимет", text):
        return number
    if re.search(r"(?:m)(?:\b|$)|метър|метра", text):
        return number * 100
    return None


def parse_attribute_measure(
    attributes: dict[str, str], names: Iterable[str], parser
) -> float | None:
    for name, value in attributes.items():
        if any(key.lower() in name.lower() for key in names):
            parsed = parser(value)
            if parsed is not None:
                return parsed
    return None


def parse_named_measure(text: str, labels: Iterable[str], parser) -> float | None:
    clean = clean_text(text)
    for label in labels:
        match = re.search(
            rf"{re.escape(label)}\s*[:\-]?\s*(\d+(?:[.,]\d+)?\s*(?:kg|mg|g|gr|гр|mm|cm|см|m)\b)",
            clean,
            flags=re.I,
        )
        if match:
            parsed = parser(match.group(1))
            if parsed is not None:
                return parsed
    return None


def parse_first_unit_measure(text: str, parser, units: str) -> float | None:
    match = re.search(rf"(\d+(?:[.,]\d+)?\s*(?:{units})(?:\b|$))", clean_text(text), flags=re.I)
    return parser(match.group(1)) if match else None


def parse_model_length_cm(title: str) -> float | None:
    # Common hard-bait model names encode length in millimetres: 130F, 125S, 90SP.
    if not re.search(r"воблер|wobbler|minnow|crankbait|jerkbait", title, flags=re.I):
        return None
    match = re.search(r"(?<!\d)(\d{2,3})(?:F|S|SP)(?:\b|\s|$)", title, flags=re.I)
    return round(float(match.group(1)) / 10, 2) if match else None


def bullets_from(product: Product) -> list[str]:
    result: list[str] = []
    for key in ("Тегло", "Дължина", "Размер", "Цвят", "Материал"):
        for attr_name, value in product.attributes.items():
            if key.lower() in attr_name.lower():
                result.append(f"{attr_name}: {value}")
                break
    if product.brand:
        result.insert(0, f"Марка: {product.brand}")
    result.append("Подходяща примамка за любителски и спортен риболов.")
    result.append("Изберете конкретната вариация преди поръчка.")
    return unique(result)[:6]


def parse_product(session: requests.Session, url: str) -> Product:
    logger.info("Product: %s", url)
    page = fetch(session, url).text
    soup = BeautifulSoup(page, "lxml")

    title_node = soup.select_one("h1.c-product-page__product-name")
    title = clean_text(title_node.get_text(" ", strip=True) if title_node else "")
    if not title:
        meta = soup.select_one('meta[property="og:title"]')
        title = clean_text(meta.get("content") if meta else "")
    if not title:
        raise ValueError(f"Missing product title: {url}")

    sku_node = soup.select_one("#ProductCode")
    base_sku = clean_text(sku_node.get_text(" ", strip=True) if sku_node else "")
    product_id_match = re.search(r"productID\s*:\s*(\d+)", page)
    product_id = product_id_match.group(1) if product_id_match else ""
    parent_sku = base_sku or product_id or re.search(r"/product/(\d+)/", url).group(1)

    description = extract_description(soup)
    attributes = parse_attributes(soup)
    all_images, image_pairs = extract_images(soup)
    if not all_images:
        og_image = soup.select_one('meta[property="og:image"]')
        if og_image and og_image.get("content"):
            all_images = [original_image_url(og_image["content"])]
            image_pairs = [(all_images[0], title)]

    groups, ordered_groups = option_groups(soup)
    variants_json = extract_json_assignment(page, "SC.ProductData.productVariants")
    base_price_node = soup.select_one('[itemprop="price"]')
    base_price = parse_decimal(base_price_node.get("content") if base_price_node and base_price_node.get("content") else (base_price_node.get_text(" ", strip=True) if base_price_node else "")) or Decimal("0")
    list_price = parse_list_price(soup, base_price)

    variants: list[Variant] = []
    for key, data in variants_json.items():
        option_values = variant_option_values(key, groups, ordered_groups)
        sku = clean_text(str(data.get("ProductVariantCode") or data.get("ProductVariantID") or parent_sku))
        parsed_variant_price = parse_decimal(data.get("ProductVariantPrice"))
        price = (
            parsed_variant_price
            if parsed_variant_price is not None and parsed_variant_price > 0
            else base_price
        )
        if price <= 0 and list_price is not None and list_price > 0:
            price = list_price
        try:
            quantity = max(0, int(float(data.get("ProductVariantQuantity", 0))))
        except (TypeError, ValueError):
            quantity = 0
        variant_weight = None
        try:
            # ProductWeight is normally in kg on this platform.
            variant_weight = float(data.get("ProductWeight")) * 1000
        except (TypeError, ValueError):
            pass
        variants.append(
            Variant(
                sku=sku,
                option_values=option_values,
                price_eur=price,
                quantity=quantity,
                weight_g=variant_weight,
                image_urls=match_variant_images(option_values, image_pairs, all_images),
            )
        )

    if not variants:
        availability = soup.select_one('meta[itemprop="availability"]')
        in_stock = bool(availability and "InStock" in str(availability.get("content", "")))
        variants = [
            Variant(
                sku=parent_sku,
                option_values={"Model": clean_text(title[-80:]) or "Standard"},
                price_eur=base_price,
                quantity=DEFAULT_IN_STOCK_QTY if in_stock else 0,
                image_urls=all_images[:10],
            )
        ]

    category_id = classify_category(title, description)
    country_origin, country_origin_source = infer_country_origin(title, description, attributes)
    if country_origin_source == "configured fallback":
        logger.warning(
            "Country of origin not published for %s; using fallback %s",
            title,
            country_origin,
        )
    product = Product(
        url=url,
        title=title,
        parent_sku=parent_sku,
        brand=infer_brand(title),
        description=description,
        bullets=[],
        category_id=category_id,
        detail_images=all_images[:100],
        attributes=attributes,
        list_price_eur=list_price,
        country_origin=country_origin,
        country_origin_source=country_origin_source,
        variants=variants,
    )
    product.bullets = bullets_from(product)
    return product


def safe_sku(value: str, fallback: str) -> str:
    value = clean_text(value) or fallback
    value = re.sub(r"[^A-Za-z0-9._-]+", "-", value).strip("-")
    return value[:80] or fallback[:80]


def variation_theme_and_cells(variant: Variant) -> tuple[str, dict[str, str]]:
    mapping = {
        "цвят": ("Color", "EU"),
        "color": ("Color", "EU"),
        "размер": ("Size", "EV"),
        "size": ("Size", "EV"),
        "модел": ("Model", "FF"),
        "model": ("Model", "FF"),
        "тегло": ("Weight", "FC"),
        "weight": ("Weight", "FC"),
        "стил": ("Style", "EW"),
        "style": ("Style", "EW"),
        "материал": ("Material", "EX"),
        "material": ("Material", "EX"),
        "аромат": ("Flavors", "EY"),
        "flavor": ("Flavors", "EY"),
        "вкус": ("Flavors", "EY"),
        "capacity": ("Capacity", "FA"),
        "капацитет": ("Capacity", "FA"),
        "quantity": ("Quantity", "FE"),
        "количество": ("Quantity", "FE"),
    }
    components: list[str] = []
    cells: dict[str, str] = {}
    for label, value in variant.option_values.items():
        normalized = label.lower().strip()
        selected = None
        for key, pair in mapping.items():
            if key in normalized:
                selected = pair
                break
        if selected is None:
            selected = ("Model", "FF")
        theme, column = selected
        if theme not in components:
            components.append(theme)
        cells[column] = value
    if not components:
        components = ["Model"]
        cells["FF"] = "Standard"
    return " × ".join(components[:2]), cells


def append_images(ws, row: int, start_col: str, images: list[str], max_images: int) -> None:
    start = ws[start_col + "2"].column
    for offset, url in enumerate(images[:max_images]):
        ws.cell(row=row, column=start + offset, value=url)


def product_dimensions(
    product: Product, variant: Variant
) -> tuple[float, float, float, float, list[str]]:
    option_text = " ".join(variant.option_values.values())

    weight = (
        parse_attribute_measure(product.attributes, ("тегло", "weight"), parse_weight_g)
        or parse_named_measure(product.description, ("тегло", "weight"), parse_weight_g)
        or parse_first_unit_measure(
            option_text, parse_weight_g, r"kg|mg|gr|g|гр|грам(?:а|ове)?"
        )
        or variant.weight_g
        or parse_first_unit_measure(
            product.title, parse_weight_g, r"kg|mg|gr|g|гр|грам(?:а|ове)?"
        )
    )
    length = (
        parse_attribute_measure(product.attributes, ("дължина", "length"), parse_length_cm)
        or parse_named_measure(product.description, ("дължина", "length"), parse_length_cm)
        or parse_first_unit_measure(option_text, parse_length_cm, r"mm|cm|см|m")
        or parse_first_unit_measure(product.title, parse_length_cm, r"mm|cm|см|m")
        or parse_model_length_cm(product.title)
    )
    width = (
        parse_attribute_measure(product.attributes, ("ширина", "width"), parse_length_cm)
        or parse_named_measure(product.description, ("ширина", "width"), parse_length_cm)
    )
    height = (
        parse_attribute_measure(
            product.attributes, ("височина", "height", "дебелина", "thickness"), parse_length_cm
        )
        or parse_named_measure(
            product.description, ("височина", "height", "дебелина", "thickness"), parse_length_cm
        )
    )

    fallbacks: list[str] = []
    if weight is None:
        weight = 10.0
        fallbacks.append("weight_g")
    if length is None:
        length = 10.0
        fallbacks.append("length_cm")
    if width is None:
        width = 1.0
        fallbacks.append("width_cm")
    if height is None:
        height = 1.0
        fallbacks.append("height_cm")

    return (
        round(weight, 2),
        round(length, 2),
        round(width, 2),
        round(height, 2),
        fallbacks,
    )


def fill_template(products: list[Product]) -> None:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")
    shutil.copy2(TEMPLATE_PATH, OUTPUT_XLSX)
    wb = load_workbook(OUTPUT_XLSX)
    ws = wb["Template"]

    # Remove old data only. Reserved/header rows 1-4 are preserved.
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    output_row = 5
    raw_rows: list[dict] = []
    out_of_stock_rows: list[dict] = []
    review_rows: list[dict] = []

    for product in products:
        mat1, mat2, mat3 = material_values(
            product.title, product.description, product.attributes, product.category_id
        )
        for variant in product.variants:
            theme, sale_cells = variation_theme_and_cells(variant)
            weight_g, length_cm, width_cm, height_cm, dimension_fallbacks = product_dimensions(
                product, variant
            )
            parent = safe_sku(product.parent_sku, f"AKVA-{output_row}")
            sku = safe_sku(variant.sku, f"{parent}-{output_row}")
            primary_image = (variant.image_urls or product.detail_images or [""])[0]

            raw_row = {
                "product_url": product.url,
                "product_name": product.title,
                "parent_sku": parent,
                "sku": sku,
                "category_id": product.category_id,
                "category_name": CATEGORY_NAMES[product.category_id],
                "brand": product.brand,
                "variation_theme": theme,
                "variation_values": json.dumps(
                    variant.option_values, ensure_ascii=False
                ),
                "quantity": variant.quantity,
                "price_eur": str(variant.price_eur),
                "list_price_eur": str(product.list_price_eur or ""),
                "weight_g": weight_g,
                "length_cm": length_cm,
                "width_cm": width_cm,
                "height_cm": height_cm,
                "main_image": primary_image,
                "all_images": " | ".join(
                    variant.image_urls or product.detail_images
                ),
                "description": product.description,
                "major_material_1": mat1,
                "major_material_2": mat2,
                "major_material_3": mat3,
                "country_origin": product.country_origin,
                "country_origin_source": product.country_origin_source,
                "dimension_fallbacks": " | ".join(dimension_fallbacks),
                "manufacturer": MANUFACTURER,
                "eu_responsible_person": EU_RESPONSIBLE_PERSON,
            }
            # The raw export always keeps every scanned variation.
            raw_rows.append(raw_row)

            # For the initial Temu upload, omit unavailable variations. A product
            # disappears completely only when all of its variations have quantity 0.
            if variant.quantity <= 0 and not INCLUDE_OUT_OF_STOCK:
                out_of_stock_rows.append(
                    {
                        "reason": "quantity_is_zero",
                        **raw_row,
                    }
                )
                continue

            # Temu requires a positive base price. Do not invent a price and do not
            # abort the whole export: place affected in-stock SKUs in a review file.
            if variant.price_eur <= 0:
                review_rows.append(
                    {
                        "reason": "missing_or_zero_price",
                        "product_url": product.url,
                        "product_name": product.title,
                        "parent_sku": parent,
                        "sku": sku,
                        "variation_values": json.dumps(
                            variant.option_values, ensure_ascii=False
                        ),
                        "quantity": variant.quantity,
                        "published_price_eur": str(variant.price_eur),
                        "category_id": product.category_id,
                        "brand": product.brand,
                        "country_origin": product.country_origin,
                        "main_image": primary_image,
                    }
                )
                logger.warning(
                    "Skipped SKU %s (%s): site published no positive price; "
                    "quantity=%s. Added to %s",
                    sku,
                    product.title,
                    variant.quantity,
                    REVIEW_CSV,
                )
                continue

            values = {
                "E": product.category_id,
                "F": CATEGORY_NAMES[product.category_id],
                "G": "Normal product",
                "L": product.title[:500],
                "M": parent,
                "N": sku,
                "O": "Add",
                "R": product.brand,
                "T": product.description[:5000],
                "DT": mat1,
                "DU": mat2,
                "DV": mat3,
                "ET": theme,
                "FH": primary_image,
                "FR": int(variant.quantity),
                "FS": float(variant.price_eur),
                "FT": product.url,
                "FW": weight_g,
                "FX": length_cm,
                "FY": width_cm,
                "FZ": height_cm,
                "GA": "Single set",
                "GB": "Yes",
                "GC": 1,
                "GD": "piece",
                "GL": SHIPPING_TEMPLATE,
                "GM": HANDLING_TIME,
                "GN": FULFILLMENT_CHANNEL,
                "GP": product.country_origin,
                "IQ": sku,
                "IR": MANUFACTURER,
                "IS": EU_RESPONSIBLE_PERSON,
            }
            if product.category_id == "32479":
                values["EL"] = "Without electricity"
            if product.list_price_eur and product.list_price_eur > variant.price_eur:
                values["FU"] = float(product.list_price_eur)
            else:
                # This is Temu's accepted alternative to a genuine list price.
                values["FV"] = "N/A"

            for col, value in sale_cells.items():
                values[col] = value
            for index, bullet in enumerate(product.bullets[:6]):
                values[get_column_letter(ws["U2"].column + index)] = bullet[:700]
            for col, value in values.items():
                ws[f"{col}{output_row}"] = value

            append_images(ws, output_row, "AA", product.detail_images, 100)
            append_images(ws, output_row, "FH", variant.image_urls or product.detail_images, 10)

            output_row += 1

    if output_row == 5:
        raise RuntimeError(
            "No active, valid SKU rows could be written. "
            "Check akvasport_out_of_stock.csv and akvasport_needs_review.csv."
        )

    validate_required_rows(ws, 5, output_row - 1)
    wb.save(OUTPUT_XLSX)
    write_csv(OUTPUT_CSV, raw_rows)
    write_csv(OUT_OF_STOCK_CSV, out_of_stock_rows)
    write_csv(REVIEW_CSV, review_rows)

    active_rows = output_row - 5
    logger.info(
        "Saved %s (%d active SKU rows; %d total scanned SKU rows)",
        OUTPUT_XLSX,
        active_rows,
        len(raw_rows),
    )
    if out_of_stock_rows:
        logger.warning(
            "%d out-of-stock SKU row(s) were excluded from the Temu workbook "
            "and saved to %s",
            len(out_of_stock_rows),
            OUT_OF_STOCK_CSV,
        )
    if review_rows:
        logger.warning(
            "%d in-stock SKU row(s) with no positive price were excluded from "
            "the Temu workbook and saved to %s",
            len(review_rows),
            REVIEW_CSV,
        )


def validate_required_rows(ws, first_row: int, last_row: int) -> None:
    errors: list[str] = []
    for row in range(first_row, last_row + 1):
        category = str(ws[f"E{row}"].value or "")
        required = REQUIRED_BY_CATEGORY.get(category, set())
        for column in required:
            value = ws[f"{column}{row}"].value
            # FU may be blank when FV explicitly contains N/A, as permitted by Temu.
            if column == "FU" and ws[f"FV{row}"].value == "N/A":
                continue
            if value is None or (isinstance(value, str) and not value.strip()):
                errors.append(f"row {row}: empty required field {column} ({ws[f'{column}2'].value})")
        if not ws[f"FH{row}"].value:
            errors.append(f"row {row}: missing SKU image")
        price = ws[f"FS{row}"].value
        if price is None or float(price) <= 0:
            errors.append(f"row {row}: invalid price")
    if errors:
        preview = "\n".join(errors[:50])
        raise ValueError(f"Template validation failed ({len(errors)} errors):\n{preview}")


def write_csv(path: Path, rows: list[dict]) -> None:
    if not rows:
        if path.exists():
            path.unlink()
        return
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    setup_logging()
    logger.info(
        "Start. Version=%s, Category=%s, product limit=%s",
        SCRAPER_VERSION,
        CATEGORY_URL,
        PRODUCT_LIMIT,
    )
    session = build_session()
    urls = collect_product_urls(session, CATEGORY_URL, PRODUCT_LIMIT)
    if not urls:
        raise RuntimeError("No product URLs were found in the selected category")
    logger.info("Found %d product URLs", len(urls))

    products: list[Product] = []
    failed: list[tuple[str, str]] = []
    for index, url in enumerate(urls, start=1):
        try:
            product = parse_product(session, url)
            products.append(product)
            logger.info(
                "Parsed %d/%d: %s (%d variants, category=%s, origin=%s [%s])",
                index,
                len(urls),
                product.title,
                len(product.variants),
                product.category_id,
                product.country_origin,
                product.country_origin_source,
            )
        except Exception as exc:  # keep the run useful if one product is malformed
            logger.exception("Failed product %s", url)
            failed.append((url, str(exc)))

    if not products:
        raise RuntimeError("All product pages failed to parse")
    fill_template(products)
    if failed:
        logger.warning("%d products failed. See log for details.", len(failed))
    logger.info("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())